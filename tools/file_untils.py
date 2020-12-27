# 文件读取
from openpyxl import load_workbook
import json
import tools.log_until as log
import configparser

def parse_json_file(filename, key):
    """读取json文件"""
    try:
        file = 'data/' + filename
        data = json.load(open(file, 'r', encoding='utf8'))
        test_data = data[key]
        log.get_log().info('open file：' + filename)
        # log.get_log().info(test_data)
        return test_data
    except Exception as error:
        log.get_log().error(error)
        return {}

def parse_excel_file(filename):
    """读取Excel文件"""
    try:
        wb = load_workbook(filename)
        ws = wb['data']
        test_data = []
        log.get_log().info('open file：' + filename)
        for x in range(2, len(tuple(ws.rows)) + 1):
            testcase_data = []
            for y in range(2, 7):
                testcase_data.append(ws.cell(row=x, column=y).value)
            test_data.append(testcase_data)
        log.get_log().info('open file：' + filename)
        return test_data
    except Exception as error:
        log.get_log().error(error)

def parse_api_file(sever, api_name):
    try:
        path = r'confglobal/api.ini'
        config = configparser.ConfigParser()
        config.read(path)
        api = config[sever][api_name]
        return api
    except Exception as error:
        log.get_log().error(error)


